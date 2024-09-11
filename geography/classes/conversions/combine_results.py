import os
import zipfile
import openpyxl
import shutil

class CombineResults:

    def __init__(self, basin_code, hyperlink_col=1):
        self.basin_code = basin_code
        self.hyperlink_col = hyperlink_col
        self.geo_download_folder = f'/Users/selenawallace/Documents/geography/data/downloads/{self.basin_code}/excel'
        self.temp_folder = f"{self.geo_download_folder}/temp_extracted/"
        if not os.path.exists(self.temp_folder):
            os.makedirs(self.temp_folder)

    # Extract ZIP files into the temporary folder
    def extract_zip_files(self):
        self.temp_folder = f"{self.geo_download_folder}temp_extracted/"
        if not os.path.exists(self.temp_folder):
            os.makedirs(self.temp_folder)

        # Loop through all ZIP files in the geo_download_folder
        zip_list = []
        for filename in os.listdir(self.geo_download_folder):
            if filename.endswith('.ZIP'):
                zip_list.append(filename)

        for filename in zip_list:
            zip_file_path = f"{self.geo_download_folder}/{filename}"
            #print(f"Extracting: {zip_file_path}") #this tells me nothing except listdir()
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
            # Extract all contents of the zip file to the temp folder
                zip_ref.extractall(self.temp_folder)

                # Here's I think the default file name
                extracted_file = "Results list for_hlead(_water_ OR river_ OR lake OR dam OR stream OR tributary OR diversion OR irrig.XLSX"
                extracted_file_path = f"{self.temp_folder}/{extracted_file}"

                new_filename = os.path.splitext(filename)[0] + '.xlsx'
                new_file_path = os.path.join(self.temp_folder, new_filename)

                # Rename the extracted file
                os.rename(extracted_file_path, new_file_path)

        print("All ZIP files have been extracted")

    # Read Excel file and extract hyperlinks
    def read_excel_with_hyperlinks(self, file_path, hyperlink_col):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        data = []
        hyperlinks = {}

        for row in ws.iter_rows(values_only=False):
            row_data = []
            for cell in row:
                if cell.column == hyperlink_col and cell.hyperlink:
                    hyperlinks[(len(data) + 1, cell.column)] = cell.hyperlink.target
                row_data.append(cell.value)
            data.append(row_data)

        return data, hyperlinks

    # Combine data from multiple files
    def combine_data_and_hyperlinks(self):
        combined_data = []
        combined_hyperlinks = {}

        for file_path in self.extracted_files:
            data, hyperlinks = self.read_excel_with_hyperlinks(file_path, self.hyperlink_col)
            start_row = len(combined_data) + 1
            combined_data.extend(data)
            for (row, col), link in hyperlinks.items():
                combined_hyperlinks[(start_row + row - 1, col)] = link

        return combined_data, combined_hyperlinks

    # Add extra columns to the combined data
    def add_additional_columns(self, data, new_columns):
        if not data or len(data) == 0:  # Check if data is empty
            print("No data to add columns to.")
            return data

        if len(data[0]) == 0:  # Check if the first row exists
            print("Row is empty.")
            return data

        # Add new columns to the header
        data[0].extend(new_columns)
        
        # Add None for each new column in every row
        for row in data[1:]:
            row.extend([None] * len(new_columns))
        
        return data

    # Write the combined data and hyperlinks to a new Excel file
    def write_combined_excel(self, data, hyperlinks, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active

        for r_idx, row in enumerate(data, start=1):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if (r_idx, c_idx) in hyperlinks:
                    cell.hyperlink = hyperlinks[(r_idx, c_idx)]

        wb.save(file_path)

    # Main method to combine everything
    def combine(self):
        self.extract_zip_files()

        # List of all extracted Excel files
        self.extracted_files = [os.path.join(self.temp_folder, f) for f in os.listdir(self.temp_folder) if f.endswith('.xlsx')]

        # Combine the data and hyperlinks
        combined_data, combined_hyperlinks = self.combine_data_and_hyperlinks()

        # Add additional columns
        new_columns = ['FileName', 'Page Numbers', 'Preliminary Review']
        combined_data = self.add_additional_columns(combined_data, new_columns)
        # need to convert the date column to date format (it's a whole mess rn)
        # and then we'll by able to sort by date before writing to excel

        # Save the combined file
        output_file = f'{self.geo_download_folder}/ResultsListCombined_{self.basin_code}_20240315.xlsx'
        self.write_combined_excel(combined_data, combined_hyperlinks, output_file)

        # Clean up temp folder
        shutil.rmtree(self.temp_folder)
        #print('Temporary folder removed')

# Example of running the class
#combine_excel = CombineResults(basin_code)
#combine_excel.combine()
