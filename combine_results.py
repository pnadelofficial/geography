import os
import zipfile
import pandas as pd
import openpyxl
from openpyxl import load_workbook
import xlsxwriter
import re
import shutil
from pathlib import Path
from datetime import datetime

#put in bcode
BCODE = 'cros'
#and make sure the paths below have a corresponding bcode folder structure

# Path to the folder containing the zip files
zip_folder = '/Users/selenawallace/Documents/Data_Science/geography2/downloads/'+ BCODE + '/excel'
output_folder = '/Users/selenawallace/Documents/Data_Science/geography2/downloads/results_to_box/' + BCODE 

# Function to extract ZIP files and move contents to output folder
def extract_zip_files(folder):
    for filename in os.listdir(folder):
        if filename.endswith('.ZIP'):
            zip_file_path = os.path.join(folder, filename)
            with zipfile.ZipFile(zip_file_path, 'r') as zip_ref:
                # Extract all files to a temporary folder
                temp_folder = os.path.join(folder, 'temp_extracted')
                zip_ref.extractall(temp_folder)
                
                # Move all files from temp folder to output folder
                for extracted_file in os.listdir(temp_folder):
                    source_file = os.path.join(temp_folder, extracted_file)
                    target_file = os.path.join(output_folder, os.path.splitext(filename)[0] + '_' + extracted_file)
                    shutil.move(source_file, target_file)
                
                # Remove temporary folder
                shutil.rmtree(temp_folder)
                print(f'Extracted: {zip_file_path}')

# Extract ZIP files and move contents to output folder
#extract_zip_files(zip_folder) 
#comment out once already done for basin

#everything above here extracts zip files to a new folder. this has been done once for bakr and doesn't need to be repeated

#list of excel files in output folder
excel_files = [os.path.join(output_folder, file) for file in os.listdir(output_folder) if file.endswith('.xlsx')]

def read_excel_with_hyperlinks(file_path, hyperlink_col=1):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    data = []
    hyperlinks = {}

    for row in ws.iter_rows(values_only=False):
        row_data = []
        for cell in row:
            if cell.column == hyperlink_col and cell.hyperlink:
                hyperlinks[(len(data) + 1, cell.column)] = cell.hyperlink.target
            row_data.append(cell.value)
        data.append(row_data)

    return data, hyperlinks

def combine_data_and_hyperlinks(file_paths, hyperlink_col=1):
    combined_data = []
    combined_hyperlinks = {}

    for file_path in file_paths:
        data, hyperlinks = read_excel_with_hyperlinks(file_path, hyperlink_col)
        start_row = len(combined_data) + 1
        combined_data.extend(data)
        for (row, col), link in hyperlinks.items():
            combined_hyperlinks[(start_row + row - 1, col)] = link

    return combined_data, combined_hyperlinks

def add_additional_columns(data, new_columns):
    data[0].extend(new_columns)
    for row in data[1:]:
        row.extend([None] * len(new_columns))
    return data

def write_combined_excel(data, hyperlinks, file_path):
    wb = openpyxl.Workbook()
    ws = wb.active

    for r_idx, row in enumerate(data, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if (r_idx, c_idx) in hyperlinks:
                cell.hyperlink = hyperlinks[(r_idx, c_idx)]

    wb.save(file_path)

# List of file paths to combine
file_paths = excel_files

# Combine the files
combined_data, combined_hyperlinks = combine_data_and_hyperlinks(file_paths, hyperlink_col=1)

# Add additional columns
new_columns = ['FileName', 'Page Numbers', 'Preliminary Review: Does the article meet the inclusion criteria? (Y/N/Needs Review)']
combined_data = add_additional_columns(combined_data, new_columns)

# Save the combined file
output_file = '/Users/selenawallace/Documents/Data_Science/geography2/downloads/results_to_box/ResultsList_' + BCODE + "_20240315.xlsx"
write_combined_excel(combined_data, combined_hyperlinks, output_file)


# combine excels into df
#combined_df = pd.concat([pd.read_excel(file) for file in excel_files])
#I think this loses hyperlinks

# Display the DataFrame with the extracted URLs
#print(combined_df)

#combined_file_path = '/Users/selenawallace/Documents/Data_Science/geography2/downloads/results_to_box/' + BCODE + '/excel_downloads.xlsx'
#combined_df.to_excel(combined_file_path, index=False)
